import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

from .models import PinsDumper, Vehicle

# PinsDumper.save() normalmente uppercasea ambiental_pin/propietary/address/
# plaque/expedition_site/model/driver vía uppercase_fields(). bulk_create()/
# bulk_update() no llaman a save() (ni señales), así que _normalize_row()
# replica ese .upper() a mano antes de armar los registros — si no, lo
# cargado por este camino quedaría en minúsculas mientras que lo creado por
# el endpoint normal (que sí pasa por save()) queda en mayúsculas.

# Todos los campos actualizables en un upsert (todo menos id/ambiental_pin,
# que es la clave de búsqueda).
UPDATABLE_FIELDS = [
    'propietary', 'address', 'phone', 'email', 'plaque', 'expedition_site',
    'model', 'capacity', 'driver', 'date_register', 'state',
]

BULK_BATCH_SIZE = 1000

# Mapeo de columnas del Excel exportado desde el PDF de la SDA
# El admin debe asegurarse que el Excel tenga estas columnas en este orden
EXPECTED_COLUMNS = [
    'PIN', 'NOMBRE', 'DIRECCION', 'TELEFONO', 'E_MAIL',
    'PLACA', 'LUGAR_EXPEDICION', 'MODELO', 'CAPACIDAD',
    'CONDUCTOR', 'FECHA_REGISTRO', 'ESTADO'
]

# Orden fijo de columnas del PDF oficial de la SDA (TRANSPORTADORES INSCRITOS
# ESCOMBROS): 12 columnas, sin excepción, en todas las filas de datos de las
# 310 páginas verificadas. Solo la página 1 trae fila de encabezado.
PDF_COLUMNS = [
    'pin', 'nombre', 'direccion', 'telefono', 'email', 'placa',
    'lugar_expedicion', 'modelo', 'capacidad', 'conductor',
    'fecha_registro', 'estado',
]


def _normalize_number(raw: str) -> str:
    """Normaliza un número que puede traer coma como separador de miles
    ("15,750" -> "15750") o como separador decimal ("12,5" -> "12.5")."""
    raw = raw.strip()
    if not raw:
        return ''
    if re.fullmatch(r'\d{1,3}(,\d{3})+', raw):
        return raw.replace(',', '')
    return raw.replace(',', '.')


def _parse_capacity(raw: str) -> Decimal | None:
    raw = _normalize_number(raw)
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _parse_phone(raw: str) -> Decimal | None:
    # El listado oficial trae teléfonos con formatos libres ("6840034 /
    # 456 34", vacíos, etc). Nos quedamos solo con los dígitos, hasta el
    # límite de max_digits=10 del modelo, y descartamos en null si no
    # queda nada aprovechable — no es un campo de identidad, no vale la
    # pena rechazar la fila completa por esto.
    digits = re.sub(r'\D', '', raw)[:10]
    if not digits:
        return None
    try:
        return Decimal(digits)
    except InvalidOperation:
        return None


def _parse_date(raw: str):
    raw = raw.strip()
    if not raw:
        return None
    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    try:
        from openpyxl.utils.datetime import from_excel
        return from_excel(int(raw)).date()
    except Exception:
        return None


def _fit(value: str, field_name: str) -> str:
    """Trunca al max_length real del campo en el modelo, para no reventar
    contra el límite de la columna en filas con texto corrupto (columnas
    del PDF que se desalinean por celdas envueltas en varias líneas)."""
    max_len = PinsDumper._meta.get_field(field_name).max_length
    return value[:max_len] if max_len else value


def _normalize_row(raw: dict, row_label) -> tuple[dict | None, str | None]:
    """Valida y normaliza una fila a un dict con los campos ya listos para
    PinsDumper (sin tocar la BD). Devuelve (record, None) o (None, motivo)
    si la fila se rechaza.
    """
    pin = raw['pin'].upper()
    pin = _fit(pin, 'ambiental_pin')
    if not pin:
        return None, f'{row_label}: PIN vacío'

    # La placa es obligatoria: sin ella el registro no sirve para nada de
    # lo que usa el sistema (imprimir el vale, enlazar Vehicle.dumper) —
    # se descarta apenas se detecta, antes de gastar tiempo parseando el
    # resto de la fila.
    plaque = _fit(raw['placa'].upper(), 'plaque')
    if not plaque:
        return None, f'{row_label} (pin={pin}): placa vacía'

    date_register = _parse_date(raw['fecha_registro'])
    if not date_register:
        return None, f'{row_label} (pin={pin}): fecha inválida "{raw["fecha_registro"]}"'

    record = {
        'ambiental_pin': pin,
        'propietary': _fit(raw['nombre'].upper(), 'propietary'),
        'address': _fit(raw['direccion'].upper(), 'address'),
        'phone': _parse_phone(raw['telefono']),
        'email': _fit(raw['email'], 'email'),
        'plaque': plaque,
        'expedition_site': _fit(raw['lugar_expedicion'].upper(), 'expedition_site'),
        'model': _fit(raw['modelo'].upper(), 'model'),
        'capacity': _parse_capacity(raw['capacidad']),
        'driver': _fit(raw['conductor'].upper(), 'driver'),
        'date_register': date_register,
        'state': raw['estado'].strip().upper() == 'ACTIVO',
    }
    return record, None


def sync_vehicle_dumper_for_plaque(plaque: str) -> int:
    """Sincroniza Vehicle.dumper con PinsDumper para UNA placa: si ya
    existe un Vehicle con esa placa, lo apunta al pin activo más reciente
    (o lo desvincula si ya ninguno quedó activo). No crea vehículos nuevos
    (requieren vehicle_type, que PinsDumper no trae). Se usa tanto al
    crear/reimportar pines como al registrar una placa nueva desde un
    viaje (RF-23): en ambos casos un Vehicle no debe quedar con dumper
    null pudiendo resolverse por placa.
    """
    if not plaque:
        return 0
    active_pin = PinsDumper.objects.filter(
        plaque=plaque, state=True
    ).order_by('-date_register').first()
    return Vehicle.objects.filter(plaque=plaque).update(dumper=active_pin)


def _sync_vehicles(touched_plaques: set) -> int:
    """Igual que sync_vehicle_dumper_for_plaque(), mas para muchas placas a
    la vez (importación masiva): primero filtra a las que ya tienen un
    Vehicle registrado en UNA consulta, en vez de dejar que cada placa
    pague su propio SELECT innecesario cuando ni siquiera hay un Vehicle
    que actualizar — con miles de filas importadas pero normalmente pocas
    placas ya registradas como Vehicle, es donde vale la pena evitarlo.
    """
    if not touched_plaques:
        return 0

    existing_plaques = set(
        Vehicle.objects.filter(plaque__in=touched_plaques).values_list('plaque', flat=True)
    )
    return sum(sync_vehicle_dumper_for_plaque(plaque) for plaque in existing_plaques)


def _plaque_winner_key(record: dict) -> tuple:
    """Orden de preferencia entre varios PIN de una misma placa: activo
    antes que finalizado (un registro vigente le gana a uno cerrado sin
    importar cuál trae fecha más reciente — se observaron ~31 placas
    donde el PIN de fecha más reciente ya está FINALIZADO mientras uno
    anterior sigue ACTIVO), y dentro del mismo estado, fecha más reciente."""
    return (record['state'], record['date_register'])


def _keep_latest_pin_per_plaque(records_by_pin: dict, row_label_by_pin: dict) -> tuple[dict, list]:
    """El listado oficial es histórico: una misma placa acumula un PIN
    distinto por cada re-registro a lo largo de los años (hasta 8 en
    algunos casos observados). Solo uno por placa es el vigente para lo
    que usa el sistema (imprimir el vale, Vehicle.dumper) — los demás se
    descartan en vez de acumular como registros muertos. Ver
    _plaque_winner_key() para el criterio de cuál se queda. Empate: gana
    el que aparece último en el archivo (mismo criterio que el dedup por
    PIN repetido)."""
    winner_pin_by_plaque: dict[str, str] = {}
    winner_key_by_plaque: dict[str, tuple] = {}

    for pin, record in records_by_pin.items():
        plaque = record['plaque']
        key = _plaque_winner_key(record)
        current_key = winner_key_by_plaque.get(plaque)
        if current_key is None or key >= current_key:
            winner_key_by_plaque[plaque] = key
            winner_pin_by_plaque[plaque] = pin

    kept = {}
    superseded = []
    for pin, record in records_by_pin.items():
        if winner_pin_by_plaque[record['plaque']] == pin:
            kept[pin] = record
        else:
            winner_pin = winner_pin_by_plaque[record['plaque']]
            superseded.append({
                'fila': row_label_by_pin[pin],
                'motivo': (
                    f'pin={pin}: histórico — la placa {record["plaque"]} tiene un '
                    f'registro más reciente (pin={winner_pin})'
                ),
            })
    return kept, superseded


def _run_import(rows: list[dict]) -> dict:
    """Normaliza todas las filas en memoria (sin tocar la BD), deduplica
    por PIN (si el archivo trae el mismo PIN repetido, gana la última
    ocurrencia — mismo criterio que un upsert fila-por-fila en orden),
    conserva solo el PIN más reciente por placa, y aplica los cambios en
    lote: UNA consulta para traer los PINS_DUMPERS ya existentes, y
    bulk_create/bulk_update para el resto, en vez de un SELECT +
    INSERT/UPDATE por cada una de las filas. Con ~16k filas esto es la
    diferencia entre ~32k round-trips a la BD y un puñado.
    """
    rejected = []
    records_by_pin: dict[str, dict] = {}
    row_label_by_pin: dict[str, str] = {}

    for row_label, raw in rows:
        record, reason = _normalize_row(raw, row_label)
        if record is None:
            rejected.append({'fila': row_label, 'motivo': reason})
            continue
        records_by_pin[record['ambiental_pin']] = record
        row_label_by_pin[record['ambiental_pin']] = row_label

    records_by_pin, superseded = _keep_latest_pin_per_plaque(records_by_pin, row_label_by_pin)
    rejected.extend(superseded)

    pins = list(records_by_pin.keys())
    existing_by_pin = {
        p.ambiental_pin: p for p in PinsDumper.objects.filter(ambiental_pin__in=pins)
    }

    to_create = []
    to_update = []
    touched_plaques: set[str] = set()

    for pin, record in records_by_pin.items():
        if record['plaque']:
            touched_plaques.add(record['plaque'])

        existing = existing_by_pin.get(pin)
        if existing:
            for field, value in record.items():
                if field == 'ambiental_pin':
                    continue
                setattr(existing, field, value)
            to_update.append(existing)
        else:
            to_create.append(PinsDumper(**record))

    try:
        with transaction.atomic():
            if to_create:
                PinsDumper.objects.bulk_create(to_create, batch_size=BULK_BATCH_SIZE)
            if to_update:
                PinsDumper.objects.bulk_update(to_update, UPDATABLE_FIELDS, batch_size=BULK_BATCH_SIZE)
    except Exception as e:
        return {
            'success': False,
            'error': f'Error al guardar en lote: {e}',
        }

    vehicles_synced = _sync_vehicles(touched_plaques)

    return {
        'success': True,
        'read': len(rows),
        'created': len(to_create),
        'updated': len(to_update),
        'rejected_count': len(rejected),
        'rejected': rejected,
        'vehicles_synced': vehicles_synced,
    }


def parse_pins_from_excel(file) -> dict:
    """
    RF-23: Lee el Excel exportado del PDF de la SDA
    y actualiza/crea registros en PinsDumper.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        if ws is None:
            return {
                'success': False,
                'error': 'El archivo Excel no tiene hojas activas.'
            }

        # Buscar la fila de encabezados
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and any(
                str(cell).upper().strip() in ['PIN', 'PLACA', 'NOMBRE']
                for cell in row if cell
            ):
                header_row = i
                break

        if not header_row:
            return {
                'success': False,
                'error': 'No se encontró la fila de encabezados. '
                         'Verifique que el Excel tenga las columnas correctas.'
            }

        # Mapear índices de columnas por nombre
        headers = [
            str(cell).upper().strip() if cell else ''
            for cell in list(ws.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True
            ))[0]
        ]

        def col_index(names: list[str]) -> int | None:
            for name in names:
                if name in headers:
                    return headers.index(name)
            return None

        idx = {
            'pin': col_index(['PIN']),
            'nombre': col_index(['NOMBRE']),
            'direccion': col_index(['DIRECCION', 'DIRECCIÓN']),
            'telefono': col_index(['TELEFONO', 'TELÉFONO']),
            'email': col_index(['E_MAIL', 'EMAIL', 'CORREO']),
            'placa': col_index(['PLACA']),
            'lugar_expedicion': col_index(['LUGAR_EXPEDICION', 'LUGAR EXPEDICIÓN', 'LUGAR_EXPEDICIÓN']),
            'modelo': col_index(['MODELO']),
            'capacidad': col_index(['CAPACIDAD', 'CAPACID']),
            'conductor': col_index(['CONDUCTOR']),
            'fecha_registro': col_index(['FECHA_REGISTRO', 'FECHA REGISTRO']),
            'estado': col_index(['ESTADO']),
        }

        if idx['pin'] is None:
            return {
                'success': False,
                'error': 'El Excel no tiene la columna PIN. '
                         'Verifique la estructura del archivo.'
            }

        def get_cell(row, key):
            i = idx[key]
            if i is None or i >= len(row):
                return ''
            val = row[i]
            return str(val).strip() if val is not None else ''

        rows = []
        for fila_num, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1
        ):
            if not any(cell for cell in row if cell):
                continue
            raw = {key: get_cell(row, key) for key in PDF_COLUMNS}
            rows.append((f'fila {fila_num}', raw))

    except Exception as e:
        return {
            'success': False,
            'error': f'No se pudo leer el archivo: {str(e)}',
        }

    return _run_import(rows)


def parse_pins_from_pdf(file) -> dict:
    """
    Lee el PDF oficial de "TRANSPORTADORES INSCRITOS ESCOMBROS" de la SDA
    (extraído con pdfplumber) y actualiza/crea registros en PinsDumper.

    El PDF trae 12 columnas fijas en todas sus filas de datos, en el mismo
    orden que PDF_COLUMNS; solo la primera página trae fila de encabezado
    (las siguientes páginas continúan la misma tabla sin repetirlo). Por
    robustez, cualquier página cuya primera fila luzca como encabezado
    también se salta.

    La extracción de tablas (page.extract_table(), CPU-bound) es el cuello
    de botella real de una importación grande — varios minutos en un PDF
    de ~300 páginas en un solo proceso — así que se reparte entre procesos
    vía pdf_extraction.extract_tables() en vez de iterar página por página
    acá mismo.
    """
    from .pdf_extraction import extract_tables

    def looks_like_header(row) -> bool:
        if not row:
            return False
        first_cell = (row[0] or '').strip().upper()
        return first_cell == 'PIN'

    try:
        tables = extract_tables(file)
        rows = []
        for page_num, table in enumerate(tables, start=1):
            if not table:
                continue
            page_rows = table
            if looks_like_header(page_rows[0]):
                page_rows = page_rows[1:]
            for row_num, row in enumerate(page_rows, start=1):
                if not row or not any((cell or '').strip() for cell in row):
                    continue
                cells = [(cell or '').strip() for cell in row]
                # Filas con más/menos columnas que las esperadas (celdas
                # envueltas en varias líneas que rompen la detección de la
                # tabla): se completan/truncan en vez de descartarse — los
                # campos de más se pierden, pero PIN/NOMBRE/DIRECCION/
                # TELEFONO al inicio de la fila siguen siendo utilizables.
                if len(cells) < len(PDF_COLUMNS):
                    cells += [''] * (len(PDF_COLUMNS) - len(cells))
                elif len(cells) > len(PDF_COLUMNS):
                    cells = cells[:len(PDF_COLUMNS)]
                raw = dict(zip(PDF_COLUMNS, cells))
                rows.append((f'página {page_num} fila {row_num}', raw))
    except Exception as e:
        return {
            'success': False,
            'error': f'No se pudo leer el PDF: {str(e)}',
        }

    if not rows:
        return {
            'success': False,
            'error': 'No se encontraron filas de datos en el PDF.',
        }

    return _run_import(rows)
